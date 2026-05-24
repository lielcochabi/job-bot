"""
Job Application Tracker — Excel file per user.
Each user (identified by NOTIFY_EMAIL) gets their own .xlsx file.
Columns: Date, Title, Company, Location, Score, Salary, URL, Method, Status, Notes

Status values: Applied | Manual - Pending | Interview | Accepted | Denied
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from job_bot.paths import TRACKER_DIR

TRACKER_DIR.mkdir(parents=True, exist_ok=True)

COLUMNS = [
    "Date Applied",
    "Job Title",
    "Company",
    "Location",
    "Match Score",
    "Salary",
    "URL",
    "Method",
    "Status",
    "Notes",
]

STATUS_OPTIONS = ["Applied", "Manual - Pending", "Interview", "Accepted", "Denied"]


def _current_user() -> str:
    """Return a stable per-user key for the tracker file.
    Priority: JOB_BOT_USERNAME env var (set by both the UI and CLI subprocesses)
    → NOTIFY_EMAIL secret → literal "default".
    """
    import os
    uname = os.environ.get("JOB_BOT_USERNAME", "")
    if uname and uname not in ("__guest__", "default", ""):
        return uname
    from secrets_manager import get_secret
    return get_secret("NOTIFY_EMAIL", "default")


def _tracker_path(user: str) -> Path:
    """Return path to the user's tracker Excel file."""
    safe = re.sub(r"[^\w]", "_", user.lower())
    return TRACKER_DIR / f"tracker_{safe}.xlsx"


def _load_workbook(path: Path):
    """Load existing workbook or create a new one with headers."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        return wb, ws

    # Create new workbook with styled header
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    # Header styling
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="1E3A5F")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # Column widths
    widths = {
        "A": 14,  # Date
        "B": 35,  # Title
        "C": 22,  # Company
        "D": 16,  # Location
        "E": 10,  # Score
        "F": 16,  # Salary
        "G": 45,  # URL
        "H": 18,  # Method
        "I": 18,  # Status
        "J": 30,  # Notes
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wb.save(path)
    return wb, ws


def _status_color(status: str) -> str:
    """Return hex fill color for a given status."""
    return {
        "Applied":          "DBEAFE",  # light blue
        "Manual - Pending": "FEF9C3",  # light yellow
        "Interview":        "D1FAE5",  # light green
        "Accepted":         "A7F3D0",  # green
        "Denied":           "FEE2E2",  # light red
    }.get(status, "FFFFFF")


def add_job(job: dict, method: str, username: str = "") -> Path:
    """
    Add a job entry to the user's tracker Excel file.
    Returns the path to the file.
    Pass username explicitly when calling from the UI; leave blank in subprocess context.
    """
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

    user = username or _current_user()
    path = _tracker_path(user)
    wb, ws   = _load_workbook(path)

    # Determine status
    is_manual = method in ("Added to manual queue", "Manual queue")
    status    = "Manual - Pending" if is_manual else "Applied"

    score = job.get("match_score") or 0
    row_data = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        job.get("title", ""),
        job.get("company", ""),
        job.get("location", ""),
        f"{score:.0f}%",
        job.get("salary", ""),
        job.get("url", ""),
        method,
        status,
        "",  # Notes — user fills this in
    ]

    # Check if job already exists (by URL) to avoid duplicates
    url = job.get("url", "")
    if url:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[6] == url:  # URL is column G (index 6)
                wb.close()
                return path

    # Styling for the new row
    row_idx    = ws.max_row + 1
    fill_color = _status_color(status)
    fill       = PatternFill("solid", fgColor=fill_color)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill   = fill
        cell.border = thin_border
        # URL as hyperlink
        if col_idx == 7 and value:
            cell.hyperlink = value
            cell.font      = Font(color="0563C1", underline="single")
            cell.alignment = left
        elif col_idx in (1, 5, 8, 9):
            cell.alignment = center
        else:
            cell.alignment = left

    ws.row_dimensions[row_idx].height = 18
    wb.save(path)
    wb.close()

    return path


def update_status(url: str, new_status: str, notes: str = "", username: str = "") -> bool:
    """
    Update the status (and optionally notes) of a job row by URL.
    Returns True if found and updated.
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    user = username or _current_user()
    path = _tracker_path(user)
    if not path.exists():
        return False

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    updated = False

    for row in ws.iter_rows(min_row=2):
        if row[6].value == url:  # URL column
            row[8].value = new_status  # Status column
            fill = PatternFill("solid", fgColor=_status_color(new_status))
            for cell in row:
                cell.fill = fill
            if notes:
                row[9].value = notes  # Notes column
            updated = True
            break

    if updated:
        wb.save(path)
    wb.close()
    return updated


def get_all_jobs(username: str = "") -> list[dict]:
    """Return all jobs from the tracker as a list of dicts."""
    user = username or _current_user()
    path = _tracker_path(user)
    if not path.exists():
        return []

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    result = []
    for row in rows:
        if not any(row):
            continue
        result.append(dict(zip(COLUMNS, row)))
    return result


def list_tracker_files() -> list[Path]:
    """List all tracker files (one per user)."""
    return sorted(TRACKER_DIR.glob("tracker_*.xlsx"))


def get_tracker_stats(jobs: list[dict]) -> dict:
    """
    Return a summary dict of application statistics from a list of tracker jobs.

    Args:
        jobs: list of job dicts as returned by get_all_jobs()

    Returns:
        dict with keys: total, by_status, response_rate, success_rate
            - total         (int)   : total number of tracked jobs
            - by_status     (dict)  : count per status string
            - response_rate (float) : % of jobs that got any response (Interview/Accepted/Denied)
            - success_rate  (float) : % of jobs that reached Interview or Accepted stage
    """
    total = len(jobs)
    if total == 0:
        return {
            "total": 0,
            "by_status": {},
            "response_rate": 0.0,
            "success_rate": 0.0,
        }

    by_status: dict[str, int] = {}
    for job in jobs:
        status = job.get("Status") or "Unknown"
        by_status[status] = by_status.get(status, 0) + 1

    responded = sum(
        by_status.get(s, 0)
        for s in ("Interview", "Accepted", "Denied")
    )
    succeeded = sum(
        by_status.get(s, 0)
        for s in ("Interview", "Accepted")
    )

    return {
        "total": total,
        "by_status": by_status,
        "response_rate": round(responded / total * 100, 1),
        "success_rate": round(succeeded / total * 100, 1),
    }
