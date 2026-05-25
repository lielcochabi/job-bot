"""
Job Bot — Streamlit Web UI
Run with: streamlit run app.py
"""
import json
import os
import re
import subprocess
import sys
import tempfile
import threading
import time
import urllib.parse
from pathlib import Path

import streamlit as st

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

# Load .env then push st.secrets — must happen before any env reads
from dotenv import load_dotenv
from job_bot.paths import DATA_DIR, ENV_PATH, JOB_CATEGORIES_PATH, ROOT

load_dotenv(ENV_PATH, override=True)
from job_bot import auth, database, secrets_manager

secrets_manager.inject_all_into_env()

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Job Bot",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded",
)

PYTHON  = str(Path(sys.executable))
BOT_DIR = ROOT

# ---------------------------------------------------------------------------
# Global CSS
# ---------------------------------------------------------------------------

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');

    /* ── Base ── */
    html, body, [class*="css"] {
        font-family: 'Inter', 'Segoe UI', system-ui, sans-serif;
    }
    .stApp {
        background: #080d18;
    }
    .block-container {
        padding-top: 2.5rem !important;
        padding-bottom: 3rem !important;
        max-width: 980px;
    }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {
        background: #0a0f1a !important;
        border-right: 1px solid #161e2e !important;
        padding: 2rem 1.2rem;
    }
    [data-testid="stSidebar"] * { color: #8b98b0 !important; }
    [data-testid="stSidebar"] hr { border-color: #161e2e !important; margin: 1.2rem 0; }

    /* active nav item */
    [data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) {
        color: #e6edf3 !important;
        background: rgba(79, 142, 247, 0.07) !important;
        border-radius: 6px !important;
        border-left: 2px solid #4f8ef7 !important;
        padding-left: 8px !important;
    }

    /* ── Page header ── */
    .page-title {
        font-size: 1.35rem; font-weight: 600;
        color: #e6edf3; margin-bottom: 0.2rem;
        letter-spacing: -0.025em;
    }
    .page-sub {
        font-size: 0.82rem; color: #3d5070;
        margin-bottom: 2rem; letter-spacing: 0.01em;
    }

    /* ── Section label ── */
    .section-label {
        font-size: 0.68rem; font-weight: 600;
        text-transform: uppercase; letter-spacing: 0.1em;
        color: #2d3d55; margin-bottom: 0.65rem;
    }
    .section-divider {
        border: none; border-top: 1px solid #111827;
        margin: 2rem 0;
    }

    /* ── Summary row ── */
    .summary-row {
        display: flex; gap: 0;
        border: 1px solid #161e2e;
        border-radius: 10px; overflow: hidden;
        margin-bottom: 2.5rem;
        background: #0c1220;
    }
    .summary-item {
        flex: 1; padding: 18px 22px;
        border-right: 1px solid #161e2e;
        transition: background 0.15s;
    }
    .summary-item:last-child { border-right: none; }
    .summary-item:hover { background: #0f1826; }
    .summary-item .num {
        font-size: 1.6rem; font-weight: 600;
        color: #e6edf3; line-height: 1;
        letter-spacing: -0.03em;
    }
    .summary-item .num.accent { color: #4f8ef7; }
    .summary-item .num.green  { color: #34d399; }
    .summary-item .lbl {
        font-size: 0.68rem; color: #3d5070;
        margin-top: 5px; text-transform: uppercase;
        letter-spacing: 0.08em;
    }

    /* ── Job card ── */
    .job-card {
        background: #0c1220;
        border: 1px solid #161e2e;
        border-radius: 10px;
        padding: 18px 22px;
        margin-bottom: 8px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.35);
        transition: border-color 0.15s, box-shadow 0.15s;
    }
    .job-card:hover {
        border-color: #2a3a55;
        box-shadow: 0 4px 16px rgba(0,0,0,0.45);
    }
    .job-card .job-title {
        font-size: 0.95rem; font-weight: 600;
        color: #dde6f0; margin-bottom: 5px;
        letter-spacing: -0.01em;
    }
    .job-card .job-meta {
        font-size: 0.78rem; color: #3d5070;
        margin-bottom: 10px; letter-spacing: 0.01em;
    }
    .score-pill {
        display: inline-block;
        padding: 2px 9px; border-radius: 4px;
        font-size: 0.72rem; font-weight: 600;
        letter-spacing: 0.02em;
    }
    .job-card .desc {
        font-size: 0.78rem; color: #3d5070;
        line-height: 1.65; margin-top: 10px;
    }
    .skill-tag {
        display: inline-block;
        background: #0f1826; color: #4f6a8a;
        border: 1px solid #1a2538;
        border-radius: 4px; padding: 1px 8px;
        font-size: 0.7rem; margin: 2px 2px 0 0;
        letter-spacing: 0.02em;
    }

    /* ── Status badge ── */
    .status-badge {
        display: inline-block; padding: 2px 8px;
        border-radius: 4px; font-size: 0.7rem;
        font-weight: 500; letter-spacing: 0.02em;
    }

    /* ── Buttons ── */
    .stButton > button {
        border-radius: 7px !important;
        font-weight: 500 !important;
        font-size: 0.83rem !important;
        padding: 0.42rem 1.1rem !important;
        min-height: 2.6rem !important;
        border: 1px solid #1e2c42 !important;
        background: #0f1826 !important;
        color: #8b98b0 !important;
        transition: background 0.15s, border-color 0.15s, color 0.15s !important;
        box-shadow: none !important;
    }
    .stButton > button:hover {
        background: #14203a !important;
        border-color: #2a3d5c !important;
        color: #c9d8e8 !important;
    }
    .stButton > button[kind="primary"] {
        background: #1a3a6e !important;
        border-color: #2a52a0 !important;
        color: #a8c4f0 !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: #1e4280 !important;
        border-color: #3a62b0 !important;
        color: #c8daf8 !important;
    }

    /* ── Task output panel — inline feed style ── */
    .task-panel-header {
        display: flex; align-items: center; gap: 8px;
        padding: 7px 4px 7px 14px;
        background: transparent;
        border-left: 2px solid transparent;
        border-bottom: 1px solid #0f1825;
        margin-top: 14px;
    }
    .task-panel-header.running { border-left-color: #f59e0b; }
    .task-panel-header.done    { border-left-color: #34d399; }
    .task-panel-header.error   { border-left-color: #f87171; }
    .task-panel-body {
        padding: 6px 0 10px 16px;
        max-height: 300px; overflow-y: auto;
        background: transparent;
        border-left: 1px solid #111827;
        margin-bottom: 6px;
    }
    /* pull the button row flush with the header, no gap */
    .task-panel-header + div [data-testid="stHorizontalBlock"] {
        margin-top: -2.1rem !important;
        justify-content: flex-end;
    }
    .task-panel-header + div [data-testid="stColumn"] {
        flex: 0 0 auto !important;
        padding: 0 2px !important;
    }
    /* tertiary button text colour in task panel */
    .task-panel-header + div button[kind="tertiary"] {
        color: #2d3d55 !important;
        font-size: 0.75rem !important;
        padding: 2px 4px !important;
    }
    .task-panel-header + div button[kind="tertiary"]:hover {
        color: #8b98b0 !important;
    }

    /* fade-out at the bottom of a long log so it doesn't hard-clip */
    .task-panel-body::after {
        content: "";
        display: block;
        height: 18px;
        background: linear-gradient(transparent, #080d18);
        pointer-events: none;
        margin-top: -18px;
        position: sticky; bottom: 0;
    }

    @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.25} }
    .running-dot {
        display: inline-block; width: 6px; height: 6px;
        border-radius: 50%; background: #f59e0b;
        animation: pulse 1.6s ease-in-out infinite;
        margin-right: 7px; flex-shrink: 0;
    }

    /* ── Inputs ── */
    [data-testid="stTextInput"] input,
    [data-testid="stTextArea"] textarea {
        background: #0c1220 !important;
        border-color: #161e2e !important;
        color: #c9d8e8 !important;
    }

    /* ── Expanders ── */
    [data-testid="stExpander"] {
        border: 1px solid #161e2e !important;
        border-radius: 9px !important;
        background: #0c1220 !important;
        margin-bottom: 10px !important;
    }

    /* ── Auth card glow ── */
    .auth-wrap {
        max-width: 440px; margin: 56px auto 0 auto;
        background: #0c1220; border: 1px solid #161e2e;
        border-radius: 14px; padding: 40px 36px;
        box-shadow: 0 0 60px rgba(79, 142, 247, 0.06),
                    0 8px 32px rgba(0,0,0,0.5);
    }

    /* ── Hide Streamlit chrome ── */
    #MainMenu, footer, header { visibility: hidden; }

    /* ── Kill the rerun dim/flash ── */
    .stApp, .stApp > *, .main, .stMain,
    [data-testid="stAppViewContainer"],
    [data-testid="stAppViewContainer"] > *,
    [data-stale="true"], [data-stale="true"] > * {
        opacity: 1 !important;
        transition: none !important;
    }
    [data-testid="stStatusWidget"] { display: none !important; }

    /* ── Step cards ── */
    .step-row {
        display: flex; gap: 10px; margin-bottom: 2rem;
    }
    .step-card {
        flex: 1; padding: 16px 18px;
        background: #0c1220; border: 1px solid #161e2e;
        border-radius: 10px; cursor: pointer;
        transition: border-color 0.15s;
        position: relative;
    }
    .step-card:hover { border-color: #2a3a55; }
    .step-card.active { border-color: #4f8ef7; }
    .step-card.done   { border-color: #1a3a2a; }
    .step-card .step-num {
        font-size: 0.65rem; font-weight: 600; letter-spacing: 0.1em;
        text-transform: uppercase; margin-bottom: 6px;
    }
    .step-card .step-title {
        font-size: 0.88rem; font-weight: 600; color: #dde6f0;
        margin-bottom: 3px;
    }
    .step-card .step-desc {
        font-size: 0.74rem; color: #3d5070; line-height: 1.4;
    }
    .step-card .step-status {
        position: absolute; top: 12px; right: 14px;
        font-size: 0.68rem; font-weight: 500;
    }

    /* ── Native bordered containers ── */
    [data-testid="stVerticalBlockBorderWrapper"] {
        background: #0c1220 !important;
        border-color: #1e2c42 !important;
        border-radius: 10px !important;
    }

    /* Dashboard quick-action row — wider gaps between buttons */
    [data-testid="stVerticalBlock"]:has(.dashboard-quick-actions-marker) + div
    [data-testid="stHorizontalBlock"] {
        gap: 1.35rem !important;
    }
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Auth initialisation & cookie check
# ---------------------------------------------------------------------------

auth.init_auth_db()

try:
    from streamlit_cookies_controller import CookieController as _CC
    _cookie_ctrl = _CC()
    _cookie_available = True
except Exception:
    _cookie_ctrl = None
    _cookie_available = False

_COOKIE_NAME = "job_bot_auth"
_COOKIE_DAYS = 90   # remember for 90 days


def _do_login(username: str, remember: bool):
    days  = _COOKIE_DAYS if remember else 1
    token = auth.create_token(username, days=days)
    st.session_state["username"]   = username
    st.session_state["auth_token"] = token
    if remember and _cookie_available:
        try:
            _cookie_ctrl.set(_COOKIE_NAME, token, max_age=days * 86400)
        except Exception:
            pass


def _do_logout():
    auth.revoke_token(st.session_state.get("auth_token", ""))
    if _cookie_available:
        try:
            _cookie_ctrl.remove(_COOKIE_NAME)
        except Exception:
            pass
    for k in ["username", "auth_token", "page_loaded"]:
        st.session_state.pop(k, None)


# ---------------------------------------------------------------------------
# Google OAuth helpers
# ---------------------------------------------------------------------------

def _google_oauth_url() -> str:
    params = {
        "client_id":     os.environ.get("GOOGLE_CLIENT_ID", ""),
        "redirect_uri":  os.environ.get("APP_URL", "http://localhost:8501"),
        "response_type": "code",
        "scope":         "openid email profile",
        "access_type":   "offline",
        "prompt":        "select_account",
    }
    return "https://accounts.google.com/o/oauth2/auth?" + urllib.parse.urlencode(params)


def _handle_google_callback() -> bool:
    """If the URL has ?code=..., exchange it for user info and log in. Returns True if handled."""
    code = st.query_params.get("code")
    if not code:
        return False

    st.query_params.clear()

    client_id     = os.environ.get("GOOGLE_CLIENT_ID", "")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET", "")
    redirect_uri  = os.environ.get("APP_URL", "http://localhost:8501")

    try:
        import httpx
        token_resp = httpx.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code":          code,
                "client_id":     client_id,
                "client_secret": client_secret,
                "redirect_uri":  redirect_uri,
                "grant_type":    "authorization_code",
            },
            timeout=10,
        )
        token_data   = token_resp.json()
        access_token = token_data.get("access_token")
        if not access_token:
            st.error("Google login failed — no access token received.")
            return True

        user_resp = httpx.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        info  = user_resp.json()
        email = info.get("email", "")
        name  = info.get("name", "")

        if not email:
            st.error("Could not retrieve your email from Google.")
            return True

        ok, result = auth.google_login(email, name)
        if ok:
            _do_login(result, remember=True)
            st.rerun()
        else:
            st.error(f"Login failed: {result}")
    except Exception as exc:
        st.error(f"Google login error: {exc}")

    return True


# Auto-login from persistent cookie
# streamlit-cookies-controller is a React component: on the very first render
# the JS hasn't mounted yet so .get() returns None even if a cookie exists.
# We do one silent rerun to let the component communicate its values, then read.
if "username" not in st.session_state and _cookie_available:
    if not st.session_state.get("_cookie_init_done"):
        # First render — component not mounted yet; rerun once to pick up cookies
        st.session_state["_cookie_init_done"] = True
        st.rerun()
    else:
        # Second render — cookies are now available
        try:
            _saved_token = _cookie_ctrl.get(_COOKIE_NAME)
            if _saved_token:
                _saved_uname = auth.validate_token(_saved_token)
                if _saved_uname:
                    st.session_state["username"]   = _saved_uname
                    st.session_state["auth_token"] = _saved_token
                else:
                    # Token expired — clear the stale cookie
                    _cookie_ctrl.remove(_COOKIE_NAME)
        except Exception:
            pass

# Handle Google OAuth redirect callback
if "username" not in st.session_state:
    _handle_google_callback()


# ---------------------------------------------------------------------------
# Auth page (shown when not logged in)
# ---------------------------------------------------------------------------

def show_auth_page():
    _, mid, _ = st.columns([1, 1.6, 1])
    with mid:
        st.markdown(
            '<div style="text-align:center;margin-bottom:1.5rem">'
            '<div style="font-size:1.4rem;font-weight:700;color:#e6edf3;letter-spacing:-0.025em">🤖 Job Bot</div>'
            '<div style="font-size:0.82rem;color:#3d5070;margin-top:4px">Automated job search — Israel & remote</div>'
            '</div>',
            unsafe_allow_html=True,
        )

        # ── Google sign-in ───────────────────────────────────────────────
        _has_google = bool(os.environ.get("GOOGLE_CLIENT_ID"))
        if _has_google:
            st.link_button(
                ":material/login: Continue with Google",
                _google_oauth_url(),
                use_container_width=True,
                type="primary",
            )
        else:
            st.button(
                ":material/login: Continue with Google",
                use_container_width=True,
                disabled=True,
                help="Add GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET to Streamlit secrets to enable Google sign-in.",
            )

        st.markdown(
            '<div style="display:flex;align-items:center;gap:8px;margin:12px 0">'
            '<hr style="flex:1;border:none;border-top:1px solid #1e2c42;margin:0">'
            '<span style="color:#3d5070;font-size:0.78rem">or</span>'
            '<hr style="flex:1;border:none;border-top:1px solid #1e2c42;margin:0">'
            '</div>',
            unsafe_allow_html=True,
        )

        tab_login, tab_signup = st.tabs([":material/login: Log in", ":material/person_add: Sign up"])

        with tab_login:
            lu  = st.text_input("Username", key="li_user", placeholder="your username")
            lp  = st.text_input("Password", type="password", key="li_pass", placeholder="••••••••")
            rem = st.checkbox("Keep me logged in", value=True, key="li_rem")
            if st.button("Log In", type="primary", use_container_width=True, key="li_btn"):
                if lu and lp:
                    ok, result = auth.login(lu, lp)
                    if ok:
                        _do_login(result, rem)
                        st.rerun()
                    else:
                        st.error(result)
                else:
                    st.warning("Please enter your username and password.")

        with tab_signup:
            su = st.text_input("Username",         key="su_user",  placeholder="choose a username")
            se = st.text_input("Email",            key="su_email", placeholder="you@example.com")
            sp = st.text_input("Password",         type="password", key="su_pass", placeholder="min 6 characters")
            sc = st.text_input("Confirm Password", type="password", key="su_conf", placeholder="repeat password")
            if st.button("Create Account", type="primary", use_container_width=True, key="su_btn"):
                if not (su and se and sp and sc):
                    st.warning("Please fill in all fields.")
                elif sp != sc:
                    st.error("Passwords don't match.")
                else:
                    ok, msg = auth.register(su, se, sp)
                    if ok:
                        ok2, uname = auth.login(su, sp)
                        if ok2:
                            _do_login(uname, remember=True)
                            st.success("Account created! Welcome.")
                            st.rerun()
                    else:
                        st.error(msg)

        # Guest access
        st.divider()
        st.caption("Just browsing? View the dashboard without an account.")
        if st.button(":material/visibility: Continue as Guest", use_container_width=True, key="guest_btn"):
            st.session_state["username"] = "__guest__"
            st.session_state["is_guest"] = True
            st.rerun()


if "username" not in st.session_state:
    show_auth_page()
    st.stop()


# ---------------------------------------------------------------------------
# Per-user context (runs for every logged-in page render)
# ---------------------------------------------------------------------------

_username = st.session_state["username"]
_is_guest = st.session_state.get("is_guest", False)

# Point the database module at this user's data
database.set_username(_username)


def _guest_block():
    """Show a login prompt and stop rendering if the user is a guest."""
    if not _is_guest:
        return
    st.markdown("""
    <div style="max-width:480px;margin:80px auto;background:#1e293b;
                border:1px solid #6366f1;border-radius:16px;
                padding:40px 32px;text-align:center;">
        <div style="font-size:1.2rem;font-weight:700;color:#e2e8f0;margin-bottom:8px">
            Login Required
        </div>
        <div style="font-size:0.9rem;color:#94a3b8;margin-bottom:24px">
            Create a free account to access this feature.
        </div>
    </div>
    """, unsafe_allow_html=True)
    _, mid, _ = st.columns([1, 1.4, 1])
    with mid:
        if st.button("Log In / Sign Up", type="primary", use_container_width=True, key="guest_login_redirect"):
            st.session_state.pop("username", None)
            st.session_state.pop("is_guest", None)
            st.rerun()
    st.stop()


# ---------------------------------------------------------------------------
# Background task helpers
# ---------------------------------------------------------------------------

# Module-level process registry — safe to access from any thread
_PROCS: dict[str, subprocess.Popen] = {}


def launch_task(cmd: list, task_key: str):
    """Start a subprocess in a background thread.
    Output goes to a temp log file; status goes to a sidecar .status file.
    The thread NEVER writes to st.session_state (not thread-safe in Streamlit).
    """
    log_path    = Path(tempfile.mktemp(suffix=".log",    prefix="jobbot_"))
    status_path = Path(str(log_path) + ".status")

    # Write initial status file
    status_path.write_text(json.dumps({"running": True, "returncode": None}))

    # Only store paths in session_state — set from main thread, never from bg thread
    st.session_state[task_key] = {
        "log_file":    str(log_path),
        "status_file": str(status_path),
        "visible":     True,
    }

    def _run():
        try:
            env = os.environ.copy()
            env["PYTHONUNBUFFERED"] = "1"
            env["JOB_BOT_USERNAME"] = _username
            with open(log_path, "w", encoding="utf-8", errors="replace") as f:
                proc = subprocess.Popen(
                    cmd,
                    stdout=f,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    cwd=str(BOT_DIR),
                    env=env,
                )
                _PROCS[task_key] = proc
                proc.wait()
            status_path.write_text(json.dumps({"running": False, "returncode": proc.returncode}))
        except Exception as exc:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"Error starting process: {exc}\n")
            status_path.write_text(json.dumps({"running": False, "returncode": 1}))
        finally:
            _PROCS.pop(task_key, None)

    threading.Thread(target=_run, daemon=True).start()


def stop_task(task_key: str, status_path: Path):
    """Kill the running process for task_key."""
    proc = _PROCS.get(task_key)
    if proc and proc.poll() is None:
        proc.terminate()
        try:
            proc.wait(timeout=3)
        except Exception:
            proc.kill()
    status_path.write_text(json.dumps({"running": False, "returncode": -1}))


_ANSI_RE = re.compile(r'\x1b\[[0-9;]*[mGKHFABCDJn]|\r')

def _format_log_line(raw: str) -> str:
    """Strip ANSI, classify the line, return a clean styled text row."""
    line = _ANSI_RE.sub("", raw).strip()
    if not line:
        return ""

    esc = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    lo  = line.lower()

    if any(k in lo for k in ("error", "fail", "exception", "traceback", "exit 1")):
        color = '#f87171'
    elif any(k in lo for k in ("[ok]", "done", "complete", "success", "added", "matched", "applied")):
        color = '#34d399'
    elif any(k in lo for k in ("warn", "skipping", "rate limit", "rate-limit", "blocked")):
        color = '#f59e0b'
    elif any(k in lo for k in ("found", "jobs found")):
        color = '#a78bfa'
    elif any(k in lo for k in ("searching", "phase")):
        color = '#4f8ef7'
    else:
        color = '#4f6a8a'

    return (
        f'<div style="padding:1px 0;line-height:1.55">'
        f'<span style="color:{color};font-size:0.79rem;'
        f'font-family:ui-monospace,\'SF Mono\',Menlo,monospace;letter-spacing:0.01em">{esc}</span>'
        f'</div>'
    )


def render_task_panel(task_key: str, title: str) -> bool:
    """
    Render the output panel for a running/completed task.
    Returns True if the task is still running (caller should rerun).
    """
    state = st.session_state.get(task_key)
    if not state or not state.get("visible", True):
        return False

    # Read running/returncode from sidecar status file (never from session_state)
    status_path = Path(state.get("status_file", ""))
    running, rc = True, None
    if status_path.exists():
        try:
            s = json.loads(status_path.read_text())
            running = s.get("running", True)
            rc      = s.get("returncode")
        except Exception:
            pass

    # Read log lines from log file
    log_path = Path(state.get("log_file", ""))
    lines: list[str] = []
    if log_path.exists():
        try:
            lines = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
            lines = [l for l in lines if l.strip()]
        except Exception:
            pass

    # ── Header row ──────────────────────────────────────────────────────────
    if running:
        _label  = f'<span class="running-dot"></span><span style="color:#e6edf3;font-size:0.85rem;font-weight:500">{title}</span><span style="color:#3d4f6b;font-size:0.78rem"> — running</span>'
        _cls    = "running"
    elif rc == 0:
        _label  = f'<span style="color:#34d399;font-size:0.85rem;font-weight:500">{title} — done ✓</span>'
        _cls    = "done"
    elif rc == -1:
        _label  = f'<span style="color:#f59e0b;font-size:0.85rem;font-weight:500">{title} — stopped</span>'
        _cls    = "error"
    else:
        _label  = f'<span style="color:#f87171;font-size:0.85rem;font-weight:500">{title} — failed (exit {rc})</span>'
        _cls    = "error"

    st.markdown(
        f'<div class="task-panel-header {_cls}">{_label}</div>',
        unsafe_allow_html=True,
    )

    # Action buttons — tertiary = no border, no background, just text
    _btn_cols = st.columns([12, 1, 1]) if running else st.columns([13, 1])
    if running:
        with _btn_cols[1]:
            if st.button("■", key=f"stop_{task_key}", type="tertiary",
                         use_container_width=True, help="Stop"):
                stop_task(task_key, status_path)
                st.rerun()
        with _btn_cols[2]:
            if st.button("✕", key=f"close_{task_key}", type="tertiary",
                         use_container_width=True, help="Dismiss"):
                st.session_state[task_key]["visible"] = False
                st.rerun()
    else:
        with _btn_cols[1]:
            if st.button("✕", key=f"close_{task_key}", type="tertiary",
                         use_container_width=True, help="Dismiss"):
                st.session_state[task_key]["visible"] = False
                st.rerun()

    # ── Output box ──────────────────────────────────────────────────────────
    if not lines:
        st.markdown(
            '<div class="task-panel-body">'
            '<span style="color:#2d3d55;font-size:0.78rem;font-style:italic">Starting…</span>'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        rows = "".join(_format_log_line(l) for l in lines[-120:] if l.strip())
        st.markdown(
            f'<div class="task-panel-body" style="padding:8px 12px">{rows}</div>',
            unsafe_allow_html=True,
        )

    return running


# ---------------------------------------------------------------------------
# Live task panel — only this fragment reruns, not the whole page
# ---------------------------------------------------------------------------

@st.fragment(run_every=0.5)
def live_task_panel(task_key: str, title: str):
    still_running = render_task_panel(task_key, title)
    if not still_running:
        # Task just finished — bust stats cache so the numbers refresh immediately
        _get_stats.clear()


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def load_config() -> dict:
    return _load_config_cached(_username)


def save_config(cfg: dict):
    database.save_config(cfg)
    _load_config_cached.clear()   # invalidate cache after save


def score_color(score: float) -> str:
    if score >= 85: return "#34d399"
    if score >= 70: return "#60a5fa"
    return "#5c6a82"


def _ai_cover_letter(job: dict, resume_text: str) -> str:
    """Generate a tailored cover letter using OpenRouter (free Llama 3 model)."""
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        return ""
    title   = job.get("title", "this role")
    company = job.get("company", "your company")
    desc    = (job.get("description") or "")[:1500]
    prompt  = (
        f"Write a concise, professional cover letter for this job application.\n\n"
        f"Resume (key highlights):\n{resume_text[:2000]}\n\n"
        f"Job: {title} at {company}\n"
        f"Description: {desc}\n\n"
        f"Instructions:\n"
        f"- 3 short paragraphs, under 220 words total\n"
        f"- Start directly (no 'Dear Hiring Manager' header)\n"
        f"- Para 1: enthusiasm + specific role/company\n"
        f"- Para 2: 2-3 most relevant skills from the resume\n"
        f"- Para 3: call to action\n"
        f"- Tone: confident, human, not generic"
    )
    try:
        import httpx as _hx
        r = _hx.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": "meta-llama/llama-3.1-8b-instruct:free",
                  "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0.7, "max_tokens": 450},
            timeout=30,
        )
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()
    except Exception as _e:
        print(f"[ai_cover_letter] {_e}")
        return ""


# ---------------------------------------------------------------------------
# ATS / form-fill helpers
# ---------------------------------------------------------------------------

# (answer_key, list-of-fragments that match a field's name/id/placeholder)
_FIELD_PATTERNS = [
    ("first_name",         ["first_name", "firstname", "fname", "first name"]),
    ("last_name",          ["last_name", "lastname", "lname", "last name", "surname"]),
    ("full_name",          ["full_name", "fullname", "your name", "applicant_name"]),
    ("email",              ["email", "e-mail", "email_address"]),
    ("phone",              ["phone", "tel", "mobile", "cell"]),
    ("linkedin_url",       ["linkedin", "linkedin_url", "linkedin_profile"]),
    ("github_url",         ["github", "github_url"]),
    ("portfolio_url",      ["portfolio", "website", "personal_site"]),
    ("location",           ["location", "city", "address", "current_location"]),
    ("cover_letter",       ["cover_letter", "coverletter", "message", "letter", "motivation", "introduction"]),
    ("years_experience",   ["years_experience", "years experience", "experience_years"]),
    ("current_title",      ["current_title", "current_position", "job_title"]),
    ("current_company",    ["current_company", "current_employer", "org"]),
    ("salary_expectation", ["salary", "compensation", "expected_salary"]),
    ("start_date",         ["start_date", "available", "availability"]),
    ("summary",            ["summary", "bio", "about_you"]),
]


def _detect_ats(url: str) -> str:
    """Identify the ATS platform from the job URL."""
    u = url.lower()
    if "boards.greenhouse.io" in u or "grnh.se" in u or "greenhouse.io" in u:
        return "greenhouse"
    if "jobs.lever.co" in u or "lever.co" in u:
        return "lever"
    if "myworkdayjobs.com" in u or "myworkday.com" in u:
        return "workday"
    if "smartrecruiters.com" in u:
        return "smartrecruiters"
    if "bamboohr.com" in u:
        return "bamboohr"
    if "ashbyhq.com" in u or "ashby.io" in u:
        return "ashby"
    return "unknown"


def _ai_form_answers(job: dict, resume_text: str) -> dict:
    """
    Generate pre-filled answers for every application form field.
    Always returns at least the user's profile data so the cheat sheet is
    useful even when the AI call fails.
    """
    cfg      = load_config()
    profile  = cfg.get("profile", {})
    name     = profile.get("name", "")
    email    = profile.get("email", os.environ.get("NOTIFY_EMAIL", ""))
    phone    = profile.get("phone", "")
    linkedin = profile.get("linkedin", "")
    location = profile.get("location", "")
    parts    = name.strip().split(None, 1) if name else ["", ""]
    fn, ln   = parts[0], (parts[1] if len(parts) > 1 else "")

    # Always-available base — populated from saved profile
    base: dict = {
        "first_name":          fn,
        "last_name":           ln,
        "full_name":           name,
        "email":               email,
        "phone":               phone,
        "linkedin_url":        linkedin,
        "location":            location,
        "country":             "Israel",
        "salary_expectation":  "According to industry standards",
        "start_date":          "Immediately",
        "work_authorization":  "Yes",
        "require_sponsorship": "No",
        "willing_to_relocate": "Yes",
        "how_did_you_hear":    "Job board",
    }

    api_key  = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        return base   # return profile data — at least the cheat sheet will have name/email

    title_   = job.get("title", "")
    company_ = job.get("company", "")
    desc_    = (job.get("description") or "")[:800]

    prompt = (
        "Fill in a job application form on behalf of the candidate below. "
        "Extract facts from the resume. Write the cover_letter specifically for this job.\n"
        "Return ONLY valid JSON — no markdown fences, no explanation.\n\n"
        f"RESUME:\n{resume_text[:2000]}\n\n"
        f"JOB: {title_} at {company_}\nDESCRIPTION: {desc_}\n\n"
        f"KNOWN: name={name!r} email={email!r} phone={phone!r} "
        f"linkedin={linkedin!r} location={location!r}\n\n"
        "Return a JSON object with keys: first_name, last_name, full_name, email, phone, "
        "linkedin_url, github_url, portfolio_url, location, city, country, current_company, "
        "current_title, years_experience, cover_letter, why_this_company, "
        "salary_expectation, start_date, work_authorization, require_sponsorship, "
        "willing_to_relocate, how_did_you_hear, summary"
    )
    try:
        import httpx as _hx
        r = _hx.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": "meta-llama/llama-3.1-8b-instruct:free",
                  "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0.3, "max_tokens": 1500},
            timeout=50,
        )
        r.raise_for_status()
        raw = r.json()["choices"][0]["message"]["content"].strip()

        ai_data: dict = {}

        # Strategy 1 — direct parse
        try:
            ai_data = json.loads(raw)
        except Exception:
            pass

        # Strategy 2 — strip ```json ... ``` fences
        if not ai_data:
            m = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', raw, re.DOTALL)
            if m:
                try:
                    ai_data = json.loads(m.group(1))
                except Exception:
                    pass

        # Strategy 3 — grab outermost { ... }
        if not ai_data:
            m = re.search(r'\{.*\}', raw, re.DOTALL)
            if m:
                try:
                    ai_data = json.loads(m.group())
                except Exception:
                    # Truncated response — attempt to close it
                    try:
                        ai_data = json.loads(m.group().rstrip(",") + '"}')
                    except Exception:
                        pass

        if ai_data:
            # Merge: AI data fills gaps; known profile values are never overwritten
            merged = {**base, **{k: v for k, v in ai_data.items() if v}}
            for k in ("email", "phone", "first_name", "last_name", "full_name", "linkedin_url", "location"):
                if base.get(k):
                    merged[k] = base[k]
            return merged

    except Exception:
        pass

    return base   # AI failed — at least return profile data


def _ensure_playwright() -> bool:
    """
    Run `playwright install chromium` unconditionally.
    Playwright is smart — it skips the download if the binary is already present,
    so calling this every time is safe and fast after the first install.
    """
    import subprocess, sys
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True, timeout=300,
        )
        return True
    except Exception:
        return False


def _playwright_fill(url: str, answers: dict, resume_bytes: bytes, resume_name: str) -> tuple:
    """
    Use a real headless Chrome browser to fill and submit a JavaScript-rendered
    application form. Works for Greenhouse, Lever, and most other ATS systems.
    """
    _ensure_playwright()
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        return "cheatsheet", "Playwright not available"

    ats = _detect_ats(url)

    # Write resume to a temp file so Playwright can upload it
    _resume_tmp = None
    if resume_bytes:
        import tempfile
        _tf = tempfile.NamedTemporaryFile(suffix=f"_{resume_name}", delete=False)
        _tf.write(resume_bytes)
        _tf.close()
        _resume_tmp = _tf.name

    def _fill(page, selector: str, value: str, timeout: int = 2000):
        try:
            page.fill(selector, value, timeout=timeout)
        except Exception:
            pass

    def _upload(page, selector: str, path: str, timeout: int = 3000):
        try:
            page.set_input_files(selector, path, timeout=timeout)
        except Exception:
            pass

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx     = browser.new_context(accept_downloads=True)
            page    = ctx.new_page()
            page.set_default_timeout(15000)

            # ── Lever ────────────────────────────────────────────────────────
            if ats == "lever":
                apply_url = url.rstrip("/")
                if not apply_url.endswith("/apply"):
                    apply_url += "/apply"
                page.goto(apply_url, wait_until="domcontentloaded", timeout=20000)
                page.wait_for_timeout(2000)

                _fill(page, '[name="name"]',     answers.get("full_name", ""))
                _fill(page, '[name="email"]',    answers.get("email", ""))
                _fill(page, '[name="phone"]',    answers.get("phone", ""))
                _fill(page, '[name="org"]',      answers.get("current_company", ""))
                _fill(page, '[name="linkedin"]', answers.get("linkedin_url", ""))
                _fill(page, '[name="github"]',   answers.get("github_url", ""))
                _fill(page, '[name="comments"]', answers.get("cover_letter", ""))
                if _resume_tmp:
                    _upload(page, 'input[type="file"]', _resume_tmp)

                page.click('[type="submit"]', timeout=5000)
                page.wait_for_timeout(3000)

            # ── Greenhouse ───────────────────────────────────────────────────
            elif ats == "greenhouse":
                page.goto(url, wait_until="domcontentloaded", timeout=20000)
                page.wait_for_timeout(2000)

                _fill(page, '#first_name', answers.get("first_name", ""))
                _fill(page, '#last_name',  answers.get("last_name", ""))
                _fill(page, '#email',      answers.get("email", ""))
                _fill(page, '#phone',      answers.get("phone", ""))
                for sel in ['#cover_letter_text', 'textarea[name*="cover"]', '#cover_letter']:
                    _fill(page, sel, answers.get("cover_letter", ""), timeout=1000)
                if _resume_tmp:
                    _upload(page, '#resume', _resume_tmp)
                    _upload(page, 'input[type="file"]', _resume_tmp)

                for btn in ['[data-qa="btn-submit"]', 'input[type="submit"]', 'button[type="submit"]']:
                    try:
                        page.click(btn, timeout=3000); break
                    except Exception:
                        pass
                page.wait_for_timeout(3000)

            # ── Generic: fill any visible form fields ────────────────────────
            else:
                page.goto(url, wait_until="domcontentloaded", timeout=20000)
                page.wait_for_timeout(2000)

                _SELECTOR_MAP = [
                    (['[name*="first"]', '[id*="first"]', '[placeholder*="First"]'],   answers.get("first_name", "")),
                    (['[name*="last"]',  '[id*="last"]',  '[placeholder*="Last"]'],    answers.get("last_name", "")),
                    (['[name*="name"]',  '[id*="name"]',  '[placeholder*="name"]'],    answers.get("full_name", "")),
                    (['[name*="email"]', '[id*="email"]', '[placeholder*="email"]'],   answers.get("email", "")),
                    (['[name*="phone"]', '[id*="phone"]', '[placeholder*="phone"]'],   answers.get("phone", "")),
                    (['[name*="linkedin"]', '[id*="linkedin"]'],                        answers.get("linkedin_url", "")),
                    (['textarea[name*="cover"]', 'textarea[id*="cover"]',
                      'textarea[name*="message"]', 'textarea[name*="letter"]'],        answers.get("cover_letter", "")),
                ]
                for selectors, value in _SELECTOR_MAP:
                    if not value:
                        continue
                    for sel in selectors:
                        try:
                            page.fill(sel, value, timeout=800); break
                        except Exception:
                            pass

                if _resume_tmp:
                    _upload(page, 'input[type="file"]', _resume_tmp)

                for btn in ['button[type="submit"]', 'input[type="submit"]', '[value="Submit"]']:
                    try:
                        page.click(btn, timeout=3000); break
                    except Exception:
                        pass
                page.wait_for_timeout(3000)

            content = page.content().lower()
            browser.close()

            if any(kw in content for kw in ("thank you", "submitted", "received your", "application sent", "success")):
                return "sent", "Application submitted via browser automation"
            return "partial", "Form filled and submitted — verify on the job site"

    except Exception as e:
        return "cheatsheet", f"Browser fill error: {e}"
    finally:
        if _resume_tmp:
            import os as _os
            try: _os.unlink(_resume_tmp)
            except Exception: pass


def _try_form_submit(url: str, answers: dict, resume_bytes: bytes, resume_name: str) -> tuple:
    """
    Try to auto-submit an online application form.
    Returns (status, message) where status is 'sent' | 'partial' | 'cheatsheet'.
    """
    import httpx as _hx
    ats = _detect_ats(url)

    # JS-heavy ATSes — skip HTTP, go straight to Playwright
    if ats in ("workday", "bamboohr", "ashby", "smartrecruiters"):
        return _playwright_fill(url, answers, resume_bytes, resume_name)

    ua = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "en-US,en;q=0.9",
    }

    try:
        from bs4 import BeautifulSoup as _BS  # noqa: N812

        fetch_url = url
        if ats == "lever" and not url.rstrip("/").endswith("/apply"):
            fetch_url = url.rstrip("/") + "/apply"

        client = _hx.Client(headers=ua, follow_redirects=True, timeout=15)
        resp0 = client.get(fetch_url)
        if resp0.status_code != 200:
            return "cheatsheet", f"Could not load form (HTTP {resp0.status_code})"

        soup = _BS(resp0.text, "html.parser")

        # Pick the application form
        if ats == "greenhouse":
            form = soup.find("form", {"id": "application_form"}) or soup.find("form")
        else:
            all_forms = soup.find_all("form")
            if not all_forms:
                # HTTP found no form — page is JS-rendered; try Playwright
                return _playwright_fill(url, answers, resume_bytes, resume_name)
            form = max(all_forms, key=lambda f: len(f.find_all(["input", "textarea", "select"])))

        if not form:
            return "cheatsheet", "No application form detected"

        action = form.get("action") or fetch_url
        method = (form.get("method") or "post").lower()
        if not action.startswith("http"):
            from urllib.parse import urljoin
            action = urljoin(fetch_url, action)

        # Seed with hidden inputs (CSRF tokens, etc.)
        data: dict = {}
        for inp in form.find_all("input", {"type": "hidden"}):
            n = inp.get("name")
            if n:
                data[n] = inp.get("value", "")

        # ATS-specific direct mappings
        if ats == "greenhouse":
            for k, v in {
                "job_application[first_name]":          answers.get("first_name", ""),
                "job_application[last_name]":           answers.get("last_name", ""),
                "job_application[email]":               answers.get("email", ""),
                "job_application[phone]":               answers.get("phone", ""),
                "job_application[cover_letter_text]":   answers.get("cover_letter", ""),
                "job_application[linkedin_profile_url]": answers.get("linkedin_url", ""),
                "job_application[website]":             answers.get("portfolio_url", ""),
            }.items():
                if v:
                    data[k] = v
        elif ats == "lever":
            for k, v in {
                "name":     answers.get("full_name", ""),
                "email":    answers.get("email", ""),
                "phone":    answers.get("phone", ""),
                "org":      answers.get("current_company", ""),
                "linkedin": answers.get("linkedin_url", ""),
                "github":   answers.get("github_url", ""),
                "comments": answers.get("cover_letter", ""),
            }.items():
                if v:
                    data[k] = v
        else:
            # Generic: pattern-match field names / ids / placeholders
            for answer_key, patterns in _FIELD_PATTERNS:
                value = answers.get(answer_key, "")
                if not value:
                    continue
                for el in form.find_all(["input", "textarea"]):
                    if el.get("type") in ("hidden", "file", "submit", "button", "checkbox", "radio"):
                        continue
                    sig = " ".join(filter(None, [
                        el.get("name", "").lower(),
                        el.get("id", "").lower(),
                        el.get("placeholder", "").lower(),
                        el.get("aria-label", "").lower(),
                    ]))
                    if any(p in sig for p in patterns):
                        n = el.get("name")
                        if n:
                            data[n] = value
                        break

        # Resume file upload
        files: dict = {}
        file_inp = form.find("input", {"type": "file"})
        if file_inp and resume_bytes:
            field_name = file_inp.get("name") or "resume"
            files[field_name] = (resume_name, resume_bytes, "application/octet-stream")

        # Submit
        if method == "post":
            resp1 = client.post(action, data=data, files=files if files else None, timeout=25)
        else:
            resp1 = client.get(action, params=data, timeout=25)

        if resp1.status_code in (200, 201, 302):
            rtext = resp1.text.lower()
            if any(kw in rtext for kw in ("thank you", "submitted", "received your", "application sent", "success")):
                return "sent", "Application submitted successfully"
            return "partial", "Form submitted — verify on the job site"
        return "cheatsheet", f"Form returned HTTP {resp1.status_code}"

    except ImportError:
        return _playwright_fill(url, answers, resume_bytes, resume_name)
    except Exception as e:
        # HTTP approach failed — fall back to Playwright
        return _playwright_fill(url, answers, resume_bytes, resume_name)


def _show_apply_cheatsheet(answers: dict, job: dict) -> None:
    """Render pre-filled application answers as a copy-paste cheat sheet."""
    jid = job.get("id", "x")
    st.markdown("##### :material/content_paste: Quick Apply Cheat Sheet")
    st.caption("Pre-filled from your resume — click any box to copy.")

    _LABELS = [
        ("first_name",         "First name"),
        ("last_name",          "Last name"),
        ("email",              "Email"),
        ("phone",              "Phone"),
        ("linkedin_url",       "LinkedIn URL"),
        ("github_url",         "GitHub URL"),
        ("portfolio_url",      "Portfolio URL"),
        ("location",           "Location / City"),
        ("current_company",    "Current company"),
        ("current_title",      "Current title"),
        ("years_experience",   "Years of experience"),
        ("salary_expectation", "Salary expectation"),
        ("start_date",         "Start date"),
        ("work_authorization", "Work authorization"),
        ("require_sponsorship","Sponsorship needed"),
        ("how_did_you_hear",   "How did you hear about us"),
    ]

    # Two-column grid for short fields
    _pairs = [(k, l, answers.get(k, "")) for k, l in _LABELS if answers.get(k, "")]
    _cols = st.columns(2)
    for i, (k, label, val) in enumerate(_pairs):
        _cols[i % 2].text_input(label, value=val, key=f"cs_{k}_{jid}")

    for key, label in (("why_this_company", "Why this company"), ("summary", "Professional summary")):
        if answers.get(key):
            st.text_area(label, value=answers[key], height=80, key=f"cs_{key}_{jid}")

    if answers.get("cover_letter"):
        st.text_area("Cover letter", value=answers["cover_letter"], height=200, key=f"cs_cl_{jid}")


def _card(job: dict, key_prefix: str = "card", *,
          show_score: bool = False, show_status: bool = False,
          skip_btn: bool = False, apply_btns: bool = False) -> None:
    """Render a single job listing as a native Streamlit bordered container."""
    score    = job.get("match_score") or 0
    title    = (job.get("title")    or "Unknown")[:80]
    company  = job.get("company")   or ""
    location = job.get("location")  or ""
    source   = job.get("source")    or ""
    url      = job.get("url")       or ""
    salary   = job.get("salary")    or ""
    desc     = (job.get("description") or "")[:280]
    status   = job.get("status")    or "found"
    jid      = job.get("id", "")

    matched_skills: list[str] = []
    try:
        matched_skills = json.loads(job.get("match_reason") or "{}").get("matched", [])
    except Exception:
        pass

    _STATUS_COLOR = {"found": "gray", "matched": "green", "skipped": "gray", "applied": "blue"}
    _STATUS_LABEL = {"found": "Unscored", "matched": "Matched", "skipped": "Skipped", "applied": "Applied"}

    with st.container(border=True):
        _ha, _hb = st.columns([5, 1])
        with _ha:
            st.markdown(f"**{title}**")
            _meta = [p for p in [company, location, source] if p]
            if salary:
                _meta.append(salary)
            if _meta:
                st.caption(" · ".join(_meta))
        with _hb:
            if show_score and score:
                if score >= 85:
                    st.markdown(f'<div style="text-align:right;color:#34d399;font-size:1.1rem;font-weight:700;padding-top:4px">{score:.0f}%</div>', unsafe_allow_html=True)
                elif score >= 70:
                    st.markdown(f'<div style="text-align:right;color:#4f8ef7;font-size:1.1rem;font-weight:700;padding-top:4px">{score:.0f}%</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="text-align:right;color:#5c6a82;font-size:1.1rem;font-weight:700;padding-top:4px">{score:.0f}%</div>', unsafe_allow_html=True)
            elif show_status:
                st.badge(
                    _STATUS_LABEL.get(status, status),
                    color=_STATUS_COLOR.get(status, "gray"),
                )

        if matched_skills:
            st.markdown("  ".join(f":blue-badge[{s}]" for s in matched_skills[:8]))

        if desc:
            st.caption(desc + ("…" if len(job.get("description") or "") > 280 else ""))

        # Action buttons
        _btn_keys: list[str] = []
        if url:          _btn_keys.append("open")
        if apply_btns:   _btn_keys += ["applied", "later", "notint"]
        elif skip_btn:   _btn_keys.append("skip")

        if _btn_keys:
            _bcols = st.columns([1] * len(_btn_keys) + [max(1, 6 - len(_btn_keys))])
            _bi = 0
            if url:
                _bcols[_bi].link_button(":material/open_in_new: Open", url, use_container_width=True)
                _bi += 1
            if apply_btns:
                if _bcols[_bi].button("Mark applied",   key=f"ma_{key_prefix}_{jid}", use_container_width=True):
                    database.set_applied(jid); st.rerun()
                _bi += 1
                if _bcols[_bi].button("Keep for later", key=f"kl_{key_prefix}_{jid}", use_container_width=True):
                    st.session_state.setdefault("snoozed_jobs", set()).add(jid); st.rerun()
                _bi += 1
                if _bcols[_bi].button("Not interested", key=f"ni_{key_prefix}_{jid}", use_container_width=True):
                    database.set_match(jid, 0, json.dumps({"reason": "Not interested"})); st.rerun()
            elif skip_btn:
                if _bcols[_bi].button("Skip", key=f"sk_{key_prefix}_{jid}", use_container_width=True):
                    database.set_match(jid, 0, json.dumps({"reason": "Manually skipped"})); st.rerun()

        # AI auto-apply (only in Apply mode)
        if apply_btns:
            _has_ai_key = bool(os.environ.get("OPENROUTER_API_KEY"))
            _col_ai, _ = st.columns([2, 5])
            if _col_ai.button(
                ":material/auto_awesome: Apply with AI",
                key=f"ai_apply_{key_prefix}_{jid}",
                use_container_width=True,
                disabled=not _has_ai_key,
                help=(
                    "Generates all form answers from your resume, then:\n"
                    "• Auto-fills Greenhouse / Lever forms\n"
                    "• Shows a Quick Apply Cheat Sheet for everything else"
                ) if _has_ai_key else "Add OPENROUTER_API_KEY to Streamlit secrets to enable.",
            ):
                _resume_text = database.get_resume_content()
                if not _resume_text:
                    st.warning(":material/info: Upload your resume in Settings first.")
                else:
                    with st.spinner("Generating application answers…"):
                        _answers = _ai_form_answers(job, _resume_text)

                    # If the full-answers call didn't produce a cover letter,
                    # fall back to the lighter cover-letter-only call
                    if not _answers.get("cover_letter"):
                        with st.spinner("Generating cover letter…"):
                            _fb_cl = _ai_cover_letter(job, _resume_text)
                        if _fb_cl:
                            _answers["cover_letter"] = _fb_cl

                    _rname, _rbytes = database.get_resume_file()
                    if not _rname:
                        _rname = "resume.pdf"

                    # Try to auto-fill ATS form (Greenhouse / Lever / generic HTML)
                    _form_status = ""
                    _form_msg    = ""
                    if url:
                        with st.spinner("Trying to auto-fill application form…"):
                            _form_status, _form_msg = _try_form_submit(url, _answers, _rbytes, _rname)

                    # Persist state for display below (survives rerun)
                    st.session_state[f"_ai_ans_{jid}"]     = _answers
                    st.session_state[f"_ai_fstatus_{jid}"] = _form_status
                    st.session_state[f"_ai_fmsg_{jid}"]    = _form_msg

                    # Only mark as applied if the form was actually submitted
                    if _form_status in ("sent", "partial"):
                        database.set_applied(jid, "AI Form")
                    st.rerun()

            # ── Show results panel ─────────────────────────────────────────────
            if f"_ai_ans_{jid}" in st.session_state:
                _ans  = st.session_state[f"_ai_ans_{jid}"]
                _fst  = st.session_state.get(f"_ai_fstatus_{jid}", "")
                _fmsg = st.session_state.get(f"_ai_fmsg_{jid}", "")

                if _fst == "sent":
                    st.success(f":material/check: {_fmsg} — click **Mark applied** to log it.")
                elif _fst == "partial":
                    st.info(f":material/info: {_fmsg} — verify on the job site, then click **Mark applied**.")
                else:
                    st.info(":material/info: Open the job link, fill the form using the answers below, "
                            "then click **Mark applied**.")
                    if _fmsg:
                        st.caption(f"Form attempt: {_fmsg}")

                _show_apply_cheatsheet(_ans, job)

# Prune stale per-job session keys so they don't accumulate forever.
# Keep only keys for jobs currently visible on this page.
_visible_jids = {j.get("id","") for j in database.get_jobs_by_status("matched")}
_stale = [k for k in list(st.session_state)
          if k.startswith(("_ai_ans_","_ai_fstatus_","_ai_fmsg_"))
          and k.split("_")[-1] not in _visible_jids]
for _k in _stale:
    st.session_state.pop(_k, None)


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------

@st.cache_resource(show_spinner=False)
def _init_db_once():
    database.init_db()
    return True

@st.cache_data(ttl=5, show_spinner=False)
def _get_stats(username: str) -> dict:
    return database.get_stats()

@st.cache_data(ttl=30, show_spinner=False)
def _has_resume_cached(username: str) -> bool:
    return database.has_resume()   # cheap count_documents, not full text load

@st.cache_data(ttl=3600, show_spinner=False)   # run at most once per hour
def _cleanup_once(username: str, expiry_days: int = 30) -> int:
    return database.cleanup_old_jobs(days=expiry_days)

@st.cache_data(ttl=60, show_spinner=False)
def _load_config_cached(username: str) -> dict:
    return database.load_config()

_init_db_once()
_startup_cfg = database.load_config()
_cleanup_once(_username, int(_startup_cfg.get("job_expiry_days", 30)))
stats = _get_stats(_username)

_PAGES = ["Dashboard", "Search", "Review matches", "Apply", "Tracker", "Settings", "Help"]

def _go(page_fragment: str):
    for p in _PAGES:
        if page_fragment in p:
            st.session_state["nav_page"]      = p
            st.session_state["_nav_from_go"]  = True  # sidebar will sync radio on next run
            break
    st.rerun()


def _go_help(section: str):
    """Open Help with a specific topic expanded."""
    st.session_state["help_section"] = section
    _go("Help")


def _help_btn(section: str, key: str):
    """Small ? button — jumps to the matching Help topic."""
    if st.button(
        ":material/help_outline:",
        key=key,
        type="tertiary",
        help="What does this mean?",
        use_container_width=True,
    ):
        _go_help(section)


if "nav_page" not in st.session_state:
    st.session_state["nav_page"] = _PAGES[0]

with st.sidebar:
    st.markdown("**:material/robot: Job Bot**")

    # If _go() was called, force the radio to the target page before it renders
    if st.session_state.pop("_nav_from_go", False):
        st.session_state["sidebar_nav"] = st.session_state["nav_page"]

    page = st.radio(
        "nav",
        _PAGES,
        label_visibility="collapsed",
        key="sidebar_nav",
    )
    st.session_state["nav_page"] = page
    st.divider()
    st.caption(
        f"{stats['total']} found · {stats['matched']} matched\n\n"
        f"{stats['applied']} applied · avg {stats['avg_score']}%"
    )
    st.divider()
    if _is_guest:
        st.caption("Browsing as guest")
        if st.button("Log in", use_container_width=True, type="primary"):
            st.session_state.pop("username", None)
            st.session_state.pop("is_guest", None)
            st.rerun()
    else:
        st.caption(f":material/person: {_username}")
        if st.button("Log out", use_container_width=True):
            _do_logout()
            st.rerun()


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

if "Dashboard" in page:
    st.title(":material/dashboard: Dashboard")
    _hint_l, _hint_r = st.columns([5, 1])
    with _hint_l:
        st.caption(
            ":material/info: Look at **Help** before starting — tap any "
            ":material/help_outline: icon next to a label to jump to that explanation."
        )
    with _hint_r:
        if st.button("Open Help", key="dash_open_help", use_container_width=True, type="secondary"):
            _go_help("overview")

    if _is_guest:
        st.info("You're browsing as a guest. Log in to use all features.")

    _has_resume  = _has_resume_cached(_username)
    _has_jobs    = stats["total"] > 0
    _has_matches = stats["matched"] > 0
    _has_ai      = bool(os.environ.get("OPENROUTER_API_KEY"))

    if not _has_resume:
        _active = 0
    elif not _has_jobs:
        _active = 1
    elif not _has_matches:
        _active = 2
    else:
        _active = 3

    def _step(i):
        if i < _active:  return "done",   "#34d399", "Done"
        if i == _active: return "active",  "#4f8ef7", "Up next"
        return "",         "#2d3d55",      ""

    s0, c0, l0 = _step(0)
    s1, c1, l1 = _step(1)
    s2, c2, l2 = _step(2)
    s3, c3, l3 = _step(3)
    _mlabel = "Score with AI" if _has_ai else "Score matches"
    _mcmd   = ["rematch", "--ai"] if _has_ai else ["rematch"]

    def _step_content(num, title, desc, s, l, help_section: str):
        _ic = (":material/check_circle:"              if s == "done"
               else ":material/radio_button_checked:" if s == "active"
               else ":material/radio_button_unchecked:")
        _bc = "green" if s == "done" else "blue" if s == "active" else "gray"
        st.caption(f"Step {num}")
        _tit, _hb = st.columns([5, 1], vertical_alignment="center")
        with _tit:
            st.markdown(f"{_ic} **{title}**")
        with _hb:
            _help_btn(help_section, key=f"help_step_{help_section}")
        st.caption(desc)
        if l:
            st.badge(l, color=_bc)

    def _step_action_btn(label, *, key, help_section, on_click=None, disabled=False):
        _l, _b, _h, _r = st.columns([0.5, 5, 0.9, 0.5], vertical_alignment="center")
        with _b:
            if st.button(label, key=key, use_container_width=True, disabled=disabled):
                if on_click:
                    on_click()
        with _h:
            _help_btn(help_section, key=f"help_act_{help_section}")

    def _launch_score():
        if _has_resume:
            launch_task([PYTHON, "-u", "main.py"] + _mcmd, "task_rematch")

    _sc = st.columns(4)

    with _sc[0].container(border=True, height="stretch"):
        _step_content(1, "Upload resume",  "Settings → Resume tab.",                               s0, l0, "upload_resume")
    with _sc[1].container(border=True, height="stretch"):
        _step_content(2, "Search jobs",    "Israeli & remote boards.",                              s1, l1, "search")
    with _sc[2].container(border=True, height="stretch"):
        _step_content(3, "Score matches",  "AI scoring (Llama 3)." if _has_ai else "Rule-based.",  s2, l2, "score_matches")
    with _sc[3].container(border=True, height="stretch"):
        _step_content(4, "Apply",          "Review and apply.",                                     s3, l3, "apply")

    with _sc[0]:
        _step_action_btn("Open Settings →", key="sc1", help_section="upload_resume", on_click=lambda: _go("Settings"))
    with _sc[1]:
        _step_action_btn("Open Search →", key="sc2", help_section="search", on_click=lambda: _go("Search"))
    with _sc[2]:
        _step_action_btn(_mlabel, key="sc3_btn", help_section="score_matches", on_click=_launch_score, disabled=not _has_jobs)
    with _sc[3]:
        _step_action_btn("Open Apply →", key="sc4", help_section="apply", on_click=lambda: _go("Apply"))

    def _metric_with_help(label: str, value, help_section: str, key: str):
        with st.container(border=True):
            _ml, _mh = st.columns([5, 1], vertical_alignment="center")
            with _ml:
                st.metric(label, value)
            with _mh:
                _help_btn(help_section, key=f"help_metric_{key}")

    with st.container(horizontal=True):
        _metric_with_help(":material/search: Found", stats["total"], "search", "found")
        _metric_with_help(":material/stars: Matched", stats["matched"], "score_matches", "matched")
        _metric_with_help(":material/send: Applied", stats["applied"], "apply", "applied")
        _metric_with_help(":material/block: Skipped", stats["skipped"], "score_matches", "skipped")
        _metric_with_help(":material/percent: Avg score", f'{stats["avg_score"]}%', "score_matches", "avg")

    if not _has_resume:
        st.caption(":material/info: Start by uploading your resume in **Settings → Resume**.")

    st.markdown('<span class="dashboard-quick-actions-marker"></span>', unsafe_allow_html=True)
    _qa1, _qa_gap, _qa2 = st.columns([1, 0.12, 1])

    def _quick_action(col, label, *, key, help_section, primary=False, disabled=False, on_click=None):
        _b, _h = col.columns([5, 1], vertical_alignment="center")
        with _b:
            if st.button(
                label,
                use_container_width=True,
                type="primary" if primary else "secondary",
                key=key,
                disabled=disabled,
            ):
                if on_click:
                    on_click()
        with _h:
            _help_btn(help_section, key=f"help_qa_{help_section}")

    _quick_action(
        _qa1, "Review matches", key="dash_qa_review", help_section="review_matches",
        disabled=not _has_jobs, on_click=lambda: _go("Review matches"),
    )
    _quick_action(
        _qa2, "Full pipeline", key="dash_qa_pipeline", help_section="full_pipeline",
        disabled=not _has_resume,
        on_click=lambda: launch_task([PYTHON, "-u", "main.py", "run", "--auto"], "task_run"),
    )

    for key, label in [("task_run", "Full Pipeline"), ("task_rematch", "Score"), ("task_search", "Search")]:
        if st.session_state.get(key):
            live_task_panel(key, label)

    with st.expander(":material/delete_sweep: Clear job database", expanded=False):
        st.caption("Remove all unmatched/unapplied jobs so you can start a fresh search for a different field.")
        if st.button(":material/delete: Clear all non-applied jobs", type="secondary"):
            _deleted = database.cleanup_old_jobs(days=0)  # days=0 wipes everything unapplied
            _get_stats.clear()          # force fresh stats on next render
            _has_resume_cached.clear()
            st.toast(f"Removed {_deleted} jobs.", icon=":material/check:")
            st.rerun()

    if _has_matches:
        st.divider()
        st.caption("Recent matches")
        for job in database.get_jobs_by_status("matched")[:5]:
            _card(job, key_prefix=f"dash_{job.get('id','')}", show_score=True, skip_btn=True)


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

elif "Search" in page:
    _guest_block()
    st.title(":material/search: Search")
    st.caption("Search job boards for new listings. Configure sources in Settings.")

    cfg = load_config()

    # ── Rate-limit alerts ───────────────────────────────────────────────────
    try:
        from datetime import datetime as _dt
        _rl = database.get_rate_limits()
        for _src, _info in _rl.items():
            try:
                _remaining = _dt.fromisoformat(_info["expires_at"]) - _dt.utcnow()
                _secs = max(0, _remaining.total_seconds())
                _hrs  = int(_secs // 3600)
                _mins = int((_secs % 3600) // 60)
                _when   = f"{_hrs}h {_mins}m" if _hrs else f"{_mins}m"
                _reason = "blocked (403)" if _info.get("hours", 1) >= 24 else "rate limited (429)"
            except Exception:
                _when, _reason = "soon", "rate limited"
            st.warning(
                f":material/block: **{_src}** is {_reason} — will retry automatically in **{_when}**."
            )
    except Exception:
        pass

    # ── Category selector ──────────────────────────────────────────────────
    _cats_data = json.loads(JOB_CATEGORIES_PATH.read_text(encoding="utf-8"))
    _saved_cats = cfg.get("selected_categories", [])

    st.caption("Categories")
    _sel_cats = st.pills(
        "categories",
        list(_cats_data.keys()),
        default=_saved_cats or [],
        selection_mode="multi",
        label_visibility="collapsed",
    )
    # Keep selection in session state only — saved to DB on search
    st.session_state["_pending_cats"] = list(_sel_cats or [])

    # Derive titles from selected categories + any custom titles saved in Settings
    _from_cats: list[str] = []
    for _c in (_sel_cats or []):
        _from_cats.extend(_cats_data.get(_c, []))
    _custom = cfg.get("custom_job_titles", [])
    _from_cats.extend(t for t in _custom if t not in _from_cats)
    all_titles = list(dict.fromkeys(_from_cats))

    # Reset title selection when categories change
    _cats_sig = "|".join(sorted(_sel_cats or []))
    if st.session_state.get("_search_cats_sig") != _cats_sig:
        st.session_state["_search_cats_sig"] = _cats_sig
        # Restore last-saved titles that are still valid for the current categories
        _last_titles = cfg.get("selected_titles", [])
        st.session_state["search_title_default"] = [t for t in _last_titles if t in all_titles]
        st.session_state["search_title_ver"] = st.session_state.get("search_title_ver", 0) + 1

    selected_titles = []

    if not all_titles:
        st.info("Select a category above to choose job titles to search for.")
    else:
        st.caption("Job titles")

        # Version counter: incrementing forces the multiselect to re-init with the new default
        if "search_title_ver" not in st.session_state:
            st.session_state["search_title_ver"] = 0
        if "search_title_default" not in st.session_state:
            st.session_state["search_title_default"] = []   # start empty
        # Drop stale entries if config changed
        st.session_state["search_title_default"] = [
            t for t in st.session_state["search_title_default"] if t in all_titles
        ]

        _tc1, _tc2, _ = st.columns([2, 2, 6])
        if _tc1.button("Select all", key="btn_sel_all", use_container_width=True):
            st.session_state["search_title_default"] = list(all_titles)
            st.session_state["search_title_ver"] += 1
            st.rerun()
        if _tc2.button("Clear all", key="btn_clr_all", use_container_width=True):
            st.session_state["search_title_default"] = []
            st.session_state["search_title_ver"] += 1
            st.rerun()

        selected_titles = st.multiselect(
            "titles",
            options=all_titles,
            default=st.session_state["search_title_default"],
            label_visibility="collapsed",
            key=f"search_ms_{st.session_state['search_title_ver']}",
        )
        # Keep in session state only — saved to DB on search
        st.session_state["search_title_default"] = list(selected_titles)
        st.session_state["_pending_titles"] = list(selected_titles)

    _loc_opts = ["Israel (on-site + remote)", "Remote only", "Worldwide"]
    if "search_loc" not in st.session_state:
        _saved = cfg.get("location_mode", "Israel (on-site + remote)")
        st.session_state["search_loc"] = _saved if _saved in _loc_opts else "Israel (on-site + remote)"

    st.caption("Location")
    _loc_mode = st.segmented_control(
        "loc_mode",
        _loc_opts,
        default=st.session_state["search_loc"],
        label_visibility="collapsed",
        key="search_loc_radio",
    )
    if _loc_mode:
        st.session_state["search_loc"] = _loc_mode

    st.divider()

    _can_search = bool(selected_titles)
    if all_titles and not selected_titles:
        st.warning("Select at least one job title.")

    if st.button("Start search", type="primary", disabled=not _can_search):
        # Persist categories + titles to DB only now (one write on search, not on every click)
        _pending_cats   = st.session_state.get("_pending_cats",   list(_sel_cats or []))
        _pending_titles = st.session_state.get("_pending_titles", list(selected_titles))
        _saved_cfg = load_config()
        _changed = (
            sorted(_pending_cats)   != sorted(_saved_cfg.get("selected_categories", [])) or
            sorted(_pending_titles) != sorted(_saved_cfg.get("selected_titles", []))
        )
        if _changed:
            _saved_cfg["selected_categories"] = _pending_cats
            _saved_cfg["selected_titles"]      = _pending_titles
            save_config(_saved_cfg)

        enabled_sources = _saved_cfg.get("selected_sources", [])
        cmd = [PYTHON, "-u", "main.py", "search"]
        for q in selected_titles:
            cmd.extend(["-q", q])
        if enabled_sources:
            cmd.extend(["--sources", ",".join(enabled_sources)])
        if _loc_mode == "Israel (on-site + remote)":
            cmd.extend(["--location", "israel"])
        elif _loc_mode == "Remote only":
            cmd.append("--remote")
        st.session_state.pop("search_result_shown", None)   # clear so results refresh after this new search
        launch_task(cmd, "task_search")

    live_task_panel("task_search", "Search")
    task_still_running = bool(st.session_state.get("task_search", {}).get("visible"))

    if not task_still_running and st.session_state.get("task_search"):
        state = st.session_state.get("task_search", {})
        status_path = Path(state.get("status_file", ""))
        if status_path.exists():
            try:
                s = json.loads(status_path.read_text())
                if not s.get("running") and not st.session_state.get("search_result_shown"):
                    st.session_state["search_result_shown"] = True
                    st.rerun()
            except Exception:
                pass

    recent = database.get_recent_jobs(hours=24)
    stats  = database.get_stats()

    if stats["total"] > 0:
        st.divider()
        st.caption(f"{stats['total']} total · {stats['found']} unscored · {stats['matched']} matched")

        if recent:
            st.caption(f":material/schedule: Found in last 24h — {len(recent)} jobs")

            search_filter = st.text_input("Filter", placeholder="Filter by title or company...", label_visibility="collapsed")

            display = recent
            if search_filter:
                fl = search_filter.lower()
                display = [j for j in recent if fl in (j.get("title") or "").lower()
                           or fl in (j.get("company") or "").lower()]

            # Split: new (found/unscored) first, already-processed below a divider
            _new_jobs  = [j for j in display[:150] if j.get("status") in ("found", "matched")]
            _done_jobs = [j for j in display[:150] if j.get("status") in ("applied", "skipped")]

            for job in _new_jobs:
                _card(job, key_prefix=f"srch_{job.get('id','')}", show_status=True)

            if _done_jobs:
                with st.expander(f":material/history: Already processed ({len(_done_jobs)})", expanded=False):
                    st.caption("These were already applied to or skipped in a previous session.")
                    for job in _done_jobs:
                        _card(job, key_prefix=f"srch_done_{job.get('id','')}", show_status=True)
        else:
            st.caption("No jobs found in the last 24 hours.")


# ---------------------------------------------------------------------------
# Review matches (before scoring)
# ---------------------------------------------------------------------------

elif "Review matches" in page:
    _guest_block()
    st.title(":material/fact_check: Review matches")
    st.caption(
        "Browse new listings from your last search. Skip roles you do not want, "
        "then score the rest on the Dashboard or below."
    )

    unscored = database.get_jobs_by_status("found")
    _has_ai = bool(os.environ.get("OPENROUTER_API_KEY"))
    _mlabel = "Score with AI" if _has_ai else "Score matches"
    _mcmd = ["rematch", "--ai"] if _has_ai else ["rematch"]

    _hdr_l, _hdr_r = st.columns([3, 1])
    with _hdr_l:
        st.metric("Awaiting review", len(unscored), border=True)
    with _hdr_r:
        if st.button(
            _mlabel,
            type="primary",
            use_container_width=True,
            disabled=not unscored or not _has_resume_cached(_username),
            key="review_page_score",
        ):
            launch_task([PYTHON, "-u", "main.py"] + _mcmd, "task_rematch")

    if st.session_state.get("task_rematch"):
        live_task_panel("task_rematch", "Score")

    if stats["total"] == 0:
        st.info("No jobs in your database yet. Run a **Search** first.")
    elif not unscored:
        st.success("All listings are scored. Open **Apply** for matches or run a new **Search**.")
        if stats.get("matched", 0) > 0:
            if st.button("Go to Apply →", type="primary", key="review_to_apply"):
                _go("Apply")
    else:
        _filter = st.text_input(
            "Filter",
            placeholder="Filter by title or company…",
            label_visibility="collapsed",
            key="review_filter",
        )
        display = unscored
        if _filter:
            fl = _filter.lower()
            display = [
                j for j in unscored
                if fl in (j.get("title") or "").lower() or fl in (j.get("company") or "").lower()
            ]

        st.caption(f"{len(display)} of {len(unscored)} unscored listing{'s' if len(unscored) != 1 else ''}")
        for job in display[:150]:
            _card(job, key_prefix=f"review_{job.get('id', '')}", show_status=True, skip_btn=True)


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

elif "Apply" in page:
    _guest_block()
    st.title(":material/send: Apply")

    matched = database.get_jobs_by_status("matched")
    snoozed = st.session_state.get("snoozed_jobs", set())
    matched = [j for j in matched if j["id"] not in snoozed]

    st.caption(f"{stats.get('applied', 0)} applied · {len(matched)} remaining")

    if not matched:
        st.info("No matched jobs to apply to. Run a search first.")
    else:
        for job in matched:
            _card(job, key_prefix="apply", show_score=True, apply_btns=True)


# ---------------------------------------------------------------------------
# Tracker
# ---------------------------------------------------------------------------

elif "Tracker" in page:
    _guest_block()

    st.title(":material/track_changes: Tracker")
    st.caption("Track outcomes for every application.")

    _T_STATUSES  = ["Applied", "Interview", "Accepted", "Denied"]
    _T_BADGE_CLR = {"Applied": "blue", "Interview": "orange", "Accepted": "green", "Denied": "red"}

    applied_jobs = database.get_applied_jobs()

    if not applied_jobs:
        st.info("No applications yet — use the Apply page to send applications.")
    else:
        from collections import Counter
        _sc = Counter(j.get("tracker_status", "Applied") for j in applied_jobs)
        with st.container(horizontal=True):
            st.metric(":material/work: Total",      len(applied_jobs),        border=True)
            st.metric(":material/send: Applied",    _sc.get("Applied", 0),    border=True)
            st.metric(":material/chat: Interview",  _sc.get("Interview", 0),  border=True)
            st.metric(":material/check: Accepted",  _sc.get("Accepted", 0),   border=True)
            st.metric(":material/close: Denied",    _sc.get("Denied", 0),     border=True)

        # ── Excel download (generated on-demand from MongoDB) ──────────────
        def _make_excel(jobs: list) -> bytes:
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            _COLS = ["Date Applied", "Job Title", "Company", "Location",
                     "Score", "URL", "Method", "Status", "Notes"]
            _CLR  = {"Applied": "DBEAFE", "Interview": "FEF9C3",
                     "Accepted": "D1FAE5", "Denied": "FEE2E2"}
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Applications"
            hfont = Font(bold=True, color="FFFFFF", size=11)
            hfill = PatternFill("solid", fgColor="1E3A5F")
            thin  = Border(*[Side(style="thin")] * 0,
                           left=Side(style="thin"), right=Side(style="thin"),
                           top=Side(style="thin"),  bottom=Side(style="thin"))
            for ci, h in enumerate(_COLS, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hfont; c.fill = hfill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin
            for ri, j in enumerate(jobs, 2):
                sc_    = j.get("match_score") or 0
                ts     = j.get("tracker_status", "Applied")
                fill_  = PatternFill("solid", fgColor=_CLR.get(ts, "FFFFFF"))
                vals   = [
                    (j.get("applied_at") or "")[:10],
                    j.get("title", ""), j.get("company", ""), j.get("location", ""),
                    f"{sc_:.0f}%" if sc_ else "",
                    j.get("url", ""), j.get("method", ""), ts,
                    j.get("tracker_notes", ""),
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.fill = fill_; c.border = thin
            buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

        st.download_button(
            ":material/download: Download Excel",
            data=_make_excel(applied_jobs),
            file_name=f"applications_{_username}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.divider()

        _tf = st.selectbox("Filter by status", ["All"] + _T_STATUSES, label_visibility="collapsed")
        _display = applied_jobs if _tf == "All" else [j for j in applied_jobs if j.get("tracker_status", "Applied") == _tf]
        st.caption(f"{len(_display)} application{'s' if len(_display) != 1 else ''}")

        for i, job in enumerate(_display):
            _jid     = job.get("id", "")
            _url     = job.get("url", "")
            _title   = job.get("title", "Unknown")
            _company = job.get("company", "")
            _score   = job.get("match_score") or 0
            _date    = (job.get("applied_at") or "")[:10]
            _method  = job.get("method", "")
            _tstatus = job.get("tracker_status", "Applied")
            _tnotes  = job.get("tracker_notes", "") or ""
            _meta    = " · ".join(p for p in [_company, _date, _method] if p)

            with st.container(border=True):
                _ca, _cb = st.columns([5, 1])
                with _ca:
                    st.markdown(f"**{_title}**")
                    if _meta:
                        st.caption(_meta)
                with _cb:
                    st.badge(_tstatus, color=_T_BADGE_CLR.get(_tstatus, "gray"))

                _cs, _cn, _csave, _copen = st.columns([2, 3, 1, 1])
                with _cs:
                    _new_status = st.selectbox(
                        "Status", _T_STATUSES,
                        index=_T_STATUSES.index(_tstatus) if _tstatus in _T_STATUSES else 0,
                        key=f"tst_{i}", label_visibility="collapsed",
                    )
                with _cn:
                    _new_notes = st.text_input(
                        "Notes", value=_tnotes,
                        key=f"tnotes_{i}", label_visibility="collapsed", placeholder="Add notes…",
                    )
                with _csave:
                    if st.button(":material/save:", key=f"tsave_{i}", use_container_width=True, help="Save"):
                        database.update_tracker_status(_jid, _new_status, _new_notes)
                        st.toast("Saved.", icon=":material/check:")
                        st.rerun()
                with _copen:
                    if _url:
                        st.link_button(":material/open_in_new:", _url, use_container_width=True)


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

elif "Settings" in page:
    _guest_block()
    st.title(":material/settings: Settings")
    st.caption("Configure your search, profile, and matching preferences.")

    cfg = load_config()

    # Save button at the top — all widget values are still collected below before saving
    _save_top = st.button("Save settings", type="primary", key="save_settings_top",
                          use_container_width=False)
    st.caption("Changes take effect after saving.")

    with st.expander(":material/description: Resume", expanded=True):
        existing_resume = database.get_resume_content()
        st.caption(f"{len(existing_resume):,} characters stored" if existing_resume else "No resume uploaded yet.")
        uploaded = st.file_uploader("Upload PDF or DOCX", type=["pdf", "docx"], label_visibility="collapsed")
        if uploaded is not None:
            from job_bot import resume_parser as rp

            suffix = Path(uploaded.name).suffix.lower()
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            resume_path = DATA_DIR / f"uploaded_resume{suffix}"
            _raw_bytes = uploaded.getvalue()
            resume_path.write_bytes(_raw_bytes)
            text = rp.extract_text(resume_path)
            if text and not text.startswith("["):
                database.save_resume(str(resume_path), text)
                database.save_resume_file(uploaded.name, _raw_bytes)   # store bytes for form uploads
                profile_data = rp.extract_resume_profile(text)
                st.session_state["auto_profile"] = profile_data
                st.toast(f"Parsed {uploaded.name} — {len(text):,} characters", icon=":material/check:")
                if profile_data.get("skills"):
                    st.markdown("  ".join(f":blue-badge[{s}]" for s in profile_data["skills"]))
            else:
                st.error(f"Could not parse {uploaded.name}.")

    with st.expander(":material/edit: Custom job titles", expanded=False):
        st.caption("Extra job titles to always include in searches, on top of whatever categories you pick. One per line.")
        custom_raw = st.text_area("Custom titles", value="\n".join(cfg.get("custom_job_titles", [])),
            height=140, label_visibility="collapsed", placeholder="e.g. Prompt Engineer\nAI Product Manager")
        custom_titles = [t.strip() for t in custom_raw.splitlines() if t.strip()]
        if custom_titles:
            st.caption(f":material/check_circle: {len(custom_titles)} custom titles active")

    with st.expander(":material/public: Job sources", expanded=False):
        st.caption("Select which job boards to search. All 8 sources are enabled by default.")
        _SOURCE_MAP = {
            "Drushim (IL)":    "drushim",
            "Indeed Israel":   "indeed_il",
            "AllJobs (IL)":    "alljobs",
            "RemoteOK":        "remoteok",
            "Remotive":        "remotive",
            "Jobicy":          "jobicy",
            "Himalayas":       "himalayas",
            "Arbeitnow":       "arbeitnow",
            "The Muse":        "themuse",
            "HN Who's Hiring": "hn",
            "Working Nomads":  "workingnomads",
        }
        _all_src_keys = list(_SOURCE_MAP.values())
        _saved_sources = cfg.get("selected_sources", _all_src_keys)
        selected_sources_settings = []
        _src_cols = st.columns(3)
        for _si, (_slabel, _skey) in enumerate(_SOURCE_MAP.items()):
            if _src_cols[_si % 3].checkbox(_slabel, value=_skey in _saved_sources, key=f"src_{_skey}"):
                selected_sources_settings.append(_skey)

    with st.expander(":material/block: Company blacklist", expanded=False):
        st.caption("Jobs from these companies will be skipped automatically.")
        blacklist_raw = st.text_area("One per line", value="\n".join(cfg.get("blacklisted_companies", [])),
            height=120, label_visibility="collapsed", placeholder="Amazon\nUber\nMeta")

    with st.expander(":material/tune: Matching", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            threshold   = st.slider("Score threshold (%)", 50, 95, int(cfg.get("match_threshold", 70)))
        with col2:
            expiry_days = st.number_input("Expire jobs after (days)", min_value=7, max_value=90, value=int(cfg.get("job_expiry_days", 30)))
        _sen_opts = ["Any", "Junior / Entry-level", "Mid-level", "Senior"]
        seniority = st.segmented_control(
            "Seniority filter",
            _sen_opts,
            default=cfg.get("seniority", "Any"),
            help="Only keep jobs that match this level. Mis-matched roles are scored 0 and skipped.",
        )
        _loc_opts_s = ["Israel (on-site + remote)", "Remote only", "Worldwide"]
        _cur_loc    = cfg.get("location_mode", "Israel (on-site + remote)")
        if _cur_loc not in _loc_opts_s:
            _cur_loc = "Israel (on-site + remote)"
        location_mode = st.segmented_control(
            "Default location for searches",
            _loc_opts_s,
            default=_cur_loc,
        )

    with st.expander(":material/person: Your profile", expanded=True):
        auto_profile = st.session_state.get("auto_profile", {})
        profile = {**cfg.get("profile", {}), **auto_profile}
        col1, col2 = st.columns(2)
        with col1:
            name     = st.text_input("Full name",  value=profile.get("name", ""))
            email    = st.text_input("Email (confirmations sent here)", value=profile.get("email", ""),
                                   placeholder="you@example.com", help="Application confirmations and BCC copies are sent to this address.")
            phone    = st.text_input("Phone",      value=profile.get("phone", ""))
        with col2:
            city     = st.text_input("City",       value=profile.get("city", ""))
            website  = st.text_input("Website",    value=profile.get("website", ""))
            linkedin = st.text_input("LinkedIn handle", value=profile.get("linkedin_handle", ""))
        summary = st.text_area("Summary", value=profile.get("summary", ""), height=100)

    st.divider()

    if st.button("Save settings", type="primary", key="save_settings_bottom") or _save_top:
        cfg["custom_job_titles"]     = custom_titles
        cfg["selected_sources"]      = selected_sources_settings
        cfg["blacklisted_companies"] = [t.strip() for t in blacklist_raw.splitlines() if t.strip()]
        cfg["location_mode"]    = location_mode
        cfg["remote_only"]      = location_mode == "Remote only"   # keep for backward compat
        cfg["match_threshold"]  = threshold
        cfg["job_expiry_days"]  = expiry_days
        cfg["seniority"]        = seniority or "Any"
        cfg["profile"] = {
            "name": name, "email": email, "phone": phone,
            "city": city, "website": website,
            "linkedin_handle": linkedin, "summary": summary,
        }
        save_config(cfg)
        st.toast("Settings saved.", icon=":material/check:")

    st.divider()
    with st.expander(":material/storage: Database", expanded=False):
        st.caption(f"Total jobs stored: {stats['total']} · Applied (kept forever): {stats['applied']}")
        col_a, col_b = st.columns(2)
        if col_a.button("Clean jobs older than 7 days", use_container_width=True):
            deleted = database.cleanup_old_jobs(days=7)
            _get_stats.clear()
            st.success(f"Removed {deleted} old jobs.")
            st.rerun()
        if col_b.button("Reset all scores for re-matching", use_container_width=True):
            count = database.reset_scores_for_rematch()
            _get_stats.clear()
            st.success(f"Reset {count} jobs.")
            st.rerun()


# ---------------------------------------------------------------------------
# Help
# ---------------------------------------------------------------------------

elif "Help" in page:
    st.title(":material/help: Help")
    st.caption("How to use Job Bot from start to finish. Use the **?** icons on the Dashboard to jump here.")

    _help_focus = st.session_state.pop("help_section", None)

    _HELP_TOPICS: list[tuple[str, str, str]] = [
        ("overview", "Getting started", """
Job Bot automates job searching in **4 steps**. Follow them in order on the Dashboard.

1. **Upload resume** — Settings → Resume  
2. **Search jobs** — fetch listings from job boards  
3. **Score matches** — rank jobs against your profile  
4. **Apply** — work through strong matches  

Tap **?** next to any Dashboard label to open the matching section below.
        """),
        ("upload_resume", "Step 1 — Upload your resume", """
Go to **Settings → Resume** and upload a PDF or DOCX.

The bot reads your skills, experience, and role title from the file and uses that
to score how well each job matches your profile. Re-upload anytime you update your CV.
        """),
        ("search", "Step 2 — Search jobs", """
Open the **Search** page, pick categories and job titles, choose location (Israel / remote / worldwide),
then click **Start search**.

The bot scans multiple job boards (RemoteOK, Remotive, Jobicy, Drushim, and more — configurable in Settings).
Results are saved to your database; running search again only **adds new** listings.

**Found** (on the Dashboard) = total jobs stored from searches.
        """),
        ("review_matches", "Optional — Review matches", """
**Not required** for the main workflow. Use **Review matches** (sidebar or Dashboard button) to browse
**unscored** listings after a search. Open postings, **Skip** roles you do not want, then run **Score matches**.

Helpful when you want to clean the list before scoring.
        """),
        ("score_matches", "Step 3 — Score matches", """
On the Dashboard, click **Score matches** or **Score with AI** (if OpenRouter is configured).

Each job gets a **0–100** match score:

- **Rule-based** — free, fast; checks skills, seniority, location, and language  
- **AI (Llama 3)** — reads the full description; needs `OPENROUTER_API_KEY`  

Jobs at or above your threshold (default **70%**) become **Matched**. Others are **Skipped**.

**Matched**, **Skipped**, and **Avg score** on the Dashboard refer to this step.
        """),
        ("apply", "Step 4 — Apply", """
Open the **Apply** page for jobs marked **Matched**.

For each listing you can:

- **Open** — view the posting  
- **Mark applied** — log it and remove from the queue  
- **Keep for later** — hide for this session  
- **Not interested** — skip permanently  
- **Apply with AI** — pre-fill forms when configured  

Track outcomes later on the **Tracker** page.
        """),
        ("full_pipeline", "Full pipeline button", """
Runs the entire flow automatically: **search → score → apply queue** (up to the apply limit).

Use when your resume and Settings are already set up. You can still review results on each page afterward.
        """),
        ("pages", "All pages", """
| Page | What it does |
|------|----------------|
| **Dashboard** | Overview, stats, and quick actions |
| **Search** | Run job searches across boards |
| **Review matches** | *(Optional)* Browse unscored jobs before scoring |
| **Apply** | Work through matched jobs |
| **Tracker** | Interview / Accepted / Denied status |
| **Settings** | Resume, sources, threshold, profile, blacklist |
| **Help** | This guide |
        """),
        ("tips", "Tips", """
- Run **Search** every few days to keep listings fresh.  
- Old unapplied jobs are removed automatically (see **Expire jobs** in Settings).  
- Lower or raise the **score threshold** in Settings if too few or too many jobs match.  
- Add companies to the **blacklist** to skip them automatically.  
        """),
    ]

    for section_id, title, body in _HELP_TOPICS:
        with st.expander(title, expanded=(section_id == (_help_focus or "overview"))):
            st.markdown(body.strip())

